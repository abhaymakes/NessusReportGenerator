import os
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
from PIL import Image, ImageDraw, ImageFont
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.drawing.image import Image as ExcelImage
import os
from datetime import datetime
import json

URL = input("Enter your nessus scan URL: ")
if os.path.exists('creds.json'):
    with open("creds.json", "r") as json_file:
        loaded_data = json.load(json_file)
        USERNAME = loaded_data.get('username')
        PASSWORD = loaded_data.get('password')

else:
    USERNAME = input('Enter nessus username: ')
    PASSWORD = input('Enter nessus password: ')
    creds = {
        'username': USERNAME,
        'password': PASSWORD
    }
    with open("creds.json", "w") as json_file:
        json.dump(creds, json_file, indent=4)

EXCEL_FILE = f"{datetime.now().strftime('%H%M%S')}.xlsx"

FONT_PATH = "dejavu-sans-mono.book.ttf"
HEADING_FONT_NAME = "Garamond"

pocs_folder = 'temp_pocs'

options = Options()
options.add_argument("user-data-dir=C:\\Abhay\\NessusGeneratorProfile")

def text_to_image(
    text,
    font_path=FONT_PATH,
    font_size=12,
    padding=14,
    bg_color="#f0f0f0",
    text_color="#4f4f4f",
    scale=2,
    image_width=1080,
):
    font_size *= scale
    padding *= scale
    image_width *= scale

    font = ImageFont.truetype(font_path, font_size)
    lines = text.splitlines() or [""]

    dummy = Image.new("RGBA", (1, 1))
    draw = ImageDraw.Draw(dummy)

    heights = [draw.textbbox((0, 0), l, font=font)[3] for l in lines]
    line_height = max(heights) + int(font_size * 0.2)

    img_height = line_height * len(lines) + padding * 2
    img = Image.new("RGBA", (image_width, img_height), bg_color)
    draw = ImageDraw.Draw(img)

    y = padding
    for line in lines:
        draw.text((padding, y), line, font=font, fill=text_color)
        y += line_height

    img = img.resize(
        (img.width // scale, img.height // scale),
        Image.Resampling.LANCZOS
    )
    return img.convert("RGB")

def setup_workbook():
    if os.path.exists(EXCEL_FILE):
        wb = load_workbook(EXCEL_FILE)
        ws_report = wb.active
        ws_report.title = "Report"
    else:
        wb = Workbook()
        ws_report = wb.active
        ws_report.title = "Report"

    ws_poc = wb["POC"] if "POC" in wb.sheetnames else wb.create_sheet("POC")

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(name=HEADING_FONT_NAME, size=9, bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers = [
        "Host", "Vulnerability Name", "Vulnerability Description",
        "Severity", "Recommendation", "Annexure"
    ]

    if ws_report.max_row == 1 and ws_report.cell(1, 1).value is None:
        for col, h in enumerate(headers, 1):
            c = ws_report.cell(1, col, h)
            c.fill = header_fill
            c.font = header_font
            c.alignment = center
        ws_report.row_dimensions[1].height = 30

    return wb, ws_report, ws_poc

def add_to_content_sheet(ws, title, image_path, row):
    header = ws.cell(row=row, column=2, value=title)
    header.font = Font(name=HEADING_FONT_NAME, size=11, bold=True)
    header.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header.fill = PatternFill("solid", fgColor="FFFF00")

    row += 2

    img = ExcelImage(image_path)
    img.anchor = f"B{row}"
    ws.add_image(img)

    rows_used = int((img.height * 0.75) / 15) + 3
    return row + rows_used

def append_to_report(ws, data, poc_row):
    base_font = Font(name=HEADING_FONT_NAME, size=8.25)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row = ws.max_row + 1
    values = [
        data["host"],
        data["title"],
        data["description"],
        data["severity"],
        data.get("solution", "N/A"),
        "POC"
    ]

    for col, val in enumerate(values, 1):
        c = ws.cell(row, col, val)
        c.font = base_font
        c.alignment = center

    colors = {
        "Critical": "8B0000",
        "High": "FF0000",
        "Medium": "FF9900",
        "Low": "FFFF00",
        "Info": "ADD8E6"
    }

    sev = ws.cell(row, 4)
    sev.fill = PatternFill("solid", fgColor=colors.get(data["severity"], "FFFFFF"))
    sev.font = Font(name=HEADING_FONT_NAME, size=8.25, bold=True, color="FFFFFF")

    annex = ws.cell(row, 6)
    annex.value = "POC"
    annex.hyperlink = f"#'POC'!B{poc_row}"
    annex.font = Font(
        name=HEADING_FONT_NAME,
        size=8.25,
        underline="single",
        color="0000FF"
    )
    annex.alignment = center

driver = webdriver.Chrome(options=options)
wait = WebDriverWait(driver, 10)
driver.get(URL)

info_count = 0
wb, ws_report, ws_poc = setup_workbook()
poc_row = 1

try:
    time.sleep(2)

    if 'Login' in driver.title:
        driver.find_element(By.CLASS_NAME, 'login-username').send_keys('admin')
        driver.find_element(By.CLASS_NAME, 'login-password').send_keys('admin')
        driver.find_element(By.CLASS_NAME, 'login-remember').click()
        driver.find_element(By.TAG_NAME, 'button').click()

    wait.until(EC.visibility_of_element_located(
        (By.CLASS_NAME, 'select2-selection__rendered'))
    ).click()

    time.sleep(1)
    driver.find_elements(By.CLASS_NAME, 'select2-results__option')[-1].click()
    time.sleep(1)

    vulns = driver.find_elements(By.CSS_SELECTOR, 'tr.vulnerability.add-plugin-id-tip')
    urls = [f"{URL}/{v.get_attribute('data-id')}" for v in vulns]

    for index, url in enumerate(urls):
        try:
            driver.get(url)
            wait.until(EC.visibility_of_element_located((By.CLASS_NAME, 'add-plugin-id-tip')))

            soup = BeautifulSoup(
                driver.find_element(By.ID, 'content').get_attribute('innerHTML'),
                'html.parser'
            )

            severity = soup.find('span', class_='severity').text.strip()
            if severity.lower() == "info":
                info_count += 1
                if info_count > 3:
                    break

            data = {
                "plugin_id": soup.find('tr', class_='noaction odd')['data-plugin-id'],
                "severity": severity,
                "title": soup.find('h4').text.strip()
            }

            details = soup.find_all('div', 'plugin-details-content')[0]
            for h, d in zip(details.find_all('h5'), details.find_all('div', 'plugin-wrap')):
                data[h.text.lower()] = d.text.strip()

            data["code_output"] = soup.find('pre', class_='monospace').text.strip()
            tds = soup.find('tr', class_='noaction odd').find_all('td')
            data["port"], data["host"] = tds[0].text.strip(), tds[1].text.strip()

            img_file = f"{data['plugin_id']}.png"
            text_to_image(data["code_output"]).save(img_file)

            poc_row = add_to_content_sheet(ws_poc, data["title"], img_file, poc_row)
            append_to_report(ws_report, data, poc_row - 3)


            wb.save(EXCEL_FILE)
            print(f"Processed: {index}. {data['title']}")

        except Exception as vuln_err:
            print(f"\n[!] Skipped vulnerability due to error: {vuln_err}")
            continue

finally:
    wb.save(EXCEL_FILE)
    driver.quit()


# import os
# import time
# import shutil
# from datetime import datetime
# from selenium import webdriver
# from selenium.webdriver.common.by import By
# from selenium.webdriver.chrome.options import Options
# from selenium.webdriver.support.wait import WebDriverWait
# from selenium.webdriver.support import expected_conditions as EC
# from bs4 import BeautifulSoup
# from PIL import Image, ImageDraw, ImageFont
# from openpyxl import Workbook, load_workbook
# from openpyxl.styles import PatternFill, Font, Alignment
# from openpyxl.drawing.image import Image as ExcelImage

# # ========================== USER INPUT ==========================
# URL = input("Enter Nessus vulnerabilities URL: ").strip()
# if not URL:
#     raise ValueError("URL cannot be empty")

# # ========================== CONSTANTS ==========================
# FONT_PATH = "dejavu-sans-mono.book.ttf"
# HEADING_FONT_NAME = "Garamond"
# POC_DIR = "pocs"

# options = Options()
# options.add_argument("user-data-dir=C:\\Abhay\\NessusGeneratorProfile")

# # ========================== UTILITIES ==========================
# def safe_mkdir(path):
#     if not os.path.exists(path):
#         os.makedirs(path)

# def progress_bar(current, total, width=40):
#     filled = int(width * current / total)
#     bar = "█" * filled + "-" * (width - filled)
#     print(f"\rProgress: |{bar}| {current}/{total}", end="", flush=True)

# # ========================== IMAGE ==========================
# def text_to_image(
#     text,
#     font_path=FONT_PATH,
#     font_size=12,
#     padding=14,
#     bg_color="#f0f0f0",
#     text_color="#4f4f4f",
#     scale=2,
#     image_width=1080,
# ):
#     font_size *= scale
#     padding *= scale
#     image_width *= scale
#     font = ImageFont.truetype(font_path, font_size)

#     lines = text.splitlines() or [""]
#     dummy = Image.new("RGBA", (1, 1))
#     draw = ImageDraw.Draw(dummy)

#     heights = [draw.textbbox((0, 0), l, font=font)[3] for l in lines]
#     line_height = max(heights) + int(font_size * 0.2)
#     img_height = line_height * len(lines) + padding * 2

#     img = Image.new("RGBA", (image_width, img_height), bg_color)
#     draw = ImageDraw.Draw(img)

#     y = padding
#     for line in lines:
#         draw.text((padding, y), line, font=font, fill=text_color)
#         y += line_height

#     img = img.resize((img.width // scale, img.height // scale), Image.Resampling.LANCZOS)
#     return img.convert("RGB")

# # ========================== EXCEL ==========================
# def setup_workbook(filename):
#     if os.path.exists(filename):
#         wb = load_workbook(filename)
#         ws_report = wb.active
#         ws_report.title = "Report"
#     else:
#         wb = Workbook()
#         ws_report = wb.active
#         ws_report.title = "Report"

#     ws_poc = wb["POC"] if "POC" in wb.sheetnames else wb.create_sheet("POC")

#     headers = [
#         "Host", "Vulnerability Name", "Vulnerability Description",
#         "Severity", "Recommendation", "Annexure"
#     ]

#     header_fill = PatternFill("solid", fgColor="1F4E78")
#     header_font = Font(name=HEADING_FONT_NAME, size=9, bold=True, color="FFFFFF")
#     center = Alignment(horizontal="center", vertical="center", wrap_text=True)

#     if ws_report.max_row == 1 and ws_report.cell(1, 1).value is None:
#         for col, h in enumerate(headers, 1):
#             c = ws_report.cell(1, col, h)
#             c.fill = header_fill
#             c.font = header_font
#             c.alignment = center
#         ws_report.row_dimensions[1].height = 30

#     return wb, ws_report, ws_poc

# def add_to_poc(ws, title, image_path, row):
#     header = ws.cell(row=row, column=2, value=title)
#     header.font = Font(name=HEADING_FONT_NAME, size=11, bold=True)
#     header.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
#     header.fill = PatternFill("solid", fgColor="FFFF00")

#     row += 2
#     img = ExcelImage(image_path)
#     img.anchor = f"B{row}"
#     ws.add_image(img)

#     rows_used = int((img.height * 0.75) / 15) + 3
#     return row + rows_used

# def append_report(ws, data, poc_row):
#     base_font = Font(name=HEADING_FONT_NAME, size=8.25)
#     center = Alignment(horizontal="center", vertical="center", wrap_text=True)

#     row = ws.max_row + 1
#     values = [
#         data["host"], data["title"], data["description"],
#         data["severity"], data.get("solution", "N/A"), "POC"
#     ]

#     for col, val in enumerate(values, 1):
#         c = ws.cell(row, col, val)
#         c.font = base_font
#         c.alignment = center

#     colors = {
#         "Critical": "8B0000",
#         "High": "FF0000",
#         "Medium": "FF9900",
#         "Low": "FFFF00",
#         "Info": "ADD8E6"
#     }

#     sev = ws.cell(row, 4)
#     sev.fill = PatternFill("solid", fgColor=colors.get(data["severity"], "FFFFFF"))
#     sev.font = Font(name=HEADING_FONT_NAME, size=8.25, bold=True, color="FFFFFF")

#     annex = ws.cell(row, 6)
#     annex.value = "POC"
#     annex.hyperlink = f"#'POC'!B{poc_row}"
#     annex.font = Font(name=HEADING_FONT_NAME, size=8.25, underline="single", color="0000FF")
#     annex.alignment = center

# # ========================== SELENIUM ==========================
# safe_mkdir(POC_DIR)

# driver = webdriver.Chrome(options=options)
# wait = WebDriverWait(driver, 10)
# driver.get(URL)

# info_count = 0
# host_for_filename = None

# try:
#     time.sleep(2)

#     if 'Login' in driver.title:
#         username_field = driver.find_element(By.CLASS_NAME, 'login-username')
#         username_field.clear()
#         username_field.send_keys('admin')

#         driver.find_element(By.CLASS_NAME, 'login-password').send_keys('admin')
#         driver.find_element(By.CLASS_NAME, 'login-remember').click()
#         driver.find_element(By.TAG_NAME, 'button').click()

#     wait.until(EC.visibility_of_element_located((By.CLASS_NAME, 'select2-selection__rendered'))).click()
#     time.sleep(1)
#     driver.find_elements(By.CLASS_NAME, 'select2-results__option')[-1].click()
#     time.sleep(1)

#     vulns = driver.find_elements(By.CSS_SELECTOR, 'tr.vulnerability.add-plugin-id-tip')
#     urls = [f"{URL}/{v.get_attribute('data-id')}" for v in vulns]

#     # Temporary workbook until host is known
#     temp_excel = "temp.xlsx"
#     wb, ws_report, ws_poc = setup_workbook(temp_excel)
#     poc_row = 1

#     total = len(urls)

#     for i, url in enumerate(urls, 1):
#         progress_bar(i, total)
#         driver.get(url)

#         soup = BeautifulSoup(driver.find_element(By.ID, 'content').get_attribute('innerHTML'), 'html.parser')
#         severity = soup.find('span', class_='severity').text.strip()

#         if severity.lower() == "info":
#             info_count += 1
#             if info_count > 3:
#                 break

#         tds = soup.find('tr', class_='noaction odd').find_all('td')
#         host = tds[1].text.strip()
#         if not host_for_filename:
#             host_for_filename = host

#         data = {
#             "plugin_id": soup.find('tr', class_='noaction odd')['data-plugin-id'],
#             "severity": severity,
#             "title": soup.find('h4').text.strip(),
#             "host": host
#         }

#         details = soup.find_all('div', 'plugin-details-content')[0]
#         for h, d in zip(details.find_all('h5'), details.find_all('div', 'plugin-wrap')):
#             data[h.text.lower()] = d.text.strip()

#         data["code_output"] = soup.find('pre', class_='monospace').text.strip()

#         img_path = os.path.join(POC_DIR, f"{data['plugin_id']}.png")
#         text_to_image(data["code_output"]).save(img_path)

#         poc_row = add_to_poc(ws_poc, data["title"], img_path, poc_row)
#         append_report(ws_report, data, poc_row - 3)

#     # ========================== FINAL SAVE ==========================
#     timestamp = datetime.now().strftime("%H:%M %d, %B")
#     final_name = f"{host_for_filename} - Scanned at {timestamp}.xlsx"
#     wb.save(final_name)

# finally:
#     driver.quit()
#     if os.path.exists(POC_DIR):
#         shutil.rmtree(POC_DIR)

# print("\nScan completed successfully.")
